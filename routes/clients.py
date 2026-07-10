# -*- coding: utf-8 -*-
# Rotas do domínio "clients" (Bloco 3 — modularização).
# Este arquivo é executado no namespace de app.py por _load_route_modules():
# tem acesso a todos os helpers/globals de app.py e registra as rotas no
# mesmo objeto Flask `app`, com URLs idênticas às originais.

@app.route('/api/clientes', methods=['GET'])
def get_clientes():
    return get_clients()


@app.route('/api/clients', methods=['GET'])
def get_clients():
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT * FROM clients ORDER BY name')
        clients = [dict_from_row(row) for row in c.fetchall()]
        conn.close()
        return jsonify(clients)
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/clients: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/cargos', methods=['GET'])
def get_positions():
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT DISTINCT position FROM clients WHERE position IS NOT NULL AND TRIM(position) != "" ORDER BY position')
        positions = [row['position'] for row in c.fetchall()]
        conn.close()
        return jsonify(positions)
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/cargos: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/empresas', methods=['GET'])
def get_companies():
    try:
        sync_accounts_from_clients()
        conn = get_db()
        c = conn.cursor()
        c.execute('''SELECT name, COALESCE(is_target, 0) as is_target FROM accounts
                     WHERE name IS NOT NULL AND TRIM(name) != ''
                     ORDER BY COALESCE(is_target, 0) DESC, name COLLATE NOCASE''')
        companies = [row['name'] for row in c.fetchall()]
        conn.close()
        return jsonify(companies)
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/empresas: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<int:client_id>', methods=['GET'])
def get_client(client_id):
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT * FROM clients WHERE id = ?', (client_id,))
        client = dict_from_row(c.fetchone())
        conn.close()
        
        if not client:
            return jsonify({'error': 'Cliente nao encontrado'}), 404
        return jsonify(client)
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/clients/{client_id}: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/check-duplicate', methods=['POST'])
def check_duplicate_client():
    try:
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        email = (data.get('email') or '').strip()
        phone = normalize_phone((data.get('phone') or '').strip())
        exclude_id = data.get('exclude_id')

        clauses = []
        params = []
        if name:
            clauses.append('LOWER(TRIM(name)) = LOWER(TRIM(?))')
            params.append(name)
        if email:
            clauses.append('LOWER(TRIM(email)) = LOWER(TRIM(?))')
            params.append(email)
        if phone:
            clauses.append('TRIM(phone) = TRIM(?)')
            params.append(phone)

        if not clauses:
            return jsonify({'matches': []})

        where = ' OR '.join(f'({c})' for c in clauses)
        sql = f'SELECT * FROM clients WHERE {where}'
        if exclude_id:
            sql += ' AND id != ?'
            params.append(int(exclude_id))

        conn = get_db()
        c = conn.cursor()
        c.execute(sql, params)
        matches = [dict_from_row(row) for row in c.fetchall()]
        conn.close()
        return jsonify({'matches': matches})
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/clients/check-duplicate: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/autofind-photo-candidates', methods=['GET'])
def clients_autofind_photo_candidates():
    try:
        name = (request.args.get('name') or '').strip()
        company = (request.args.get('company') or '').strip()
        logger.info(f'[AutoPic] GET /autofind-photo-candidates: name={name!r} company={company!r}')

        if not name and not company:
            logger.warning('[AutoPic] GET /autofind-photo-candidates: nome e empresa vazios, abortando.')
            return jsonify({'error': 'Informe ao menos nome ou empresa.'}), 400

        candidates = []
        queries_tried = []

        # Enriquecimento via LLM: descobrir o cargo da pessoa — limitado a 10s para não travar a UI.
        # O _sai_simple_prompt tem timeout de 45s; usamos concurrent.futures para encurtar.
        role = None
        if name and company:
            try:
                with concurrent.futures.ThreadPoolExecutor(max_workers=1) as _ex:
                    _fut = _ex.submit(_autopic_get_role_via_llm, name, company)
                    try:
                        role = _fut.result(timeout=10)
                    except concurrent.futures.TimeoutError:
                        logger.warning(f'[AutoPic] Timeout 10s ao buscar cargo via LLM para {name!r}, prosseguindo sem cargo.')
            except Exception as e:
                logger.warning(f'[AutoPic] Falha ao obter cargo via LLM: {e}')

        def _merge_candidates(existing, new_urls, limit=6):
            seen = set(existing)
            result = list(existing)
            for u in new_urls:
                if u not in seen:
                    seen.add(u)
                    result.append(u)
            return result[:limit]

        # Estratégia 1 (primária): nome ENTRE ASPAS + empresa sem aspas.
        # As aspas no nome são essenciais: sem elas "Henrique Netto" vira dois tokens separados
        # e o Bing pode combinar com a dupla sertaneja "Netto & Henrique" ou similares.
        # Com aspas, o Bing exige a frase exata e prioriza o profissional da empresa.
        if name and company:
            q1 = f'"{name}" {company}'
            queries_tried.append(q1)
            logger.info(f'[AutoPic] Estratégia 1 (nome entre aspas + empresa): query={q1!r}')
            candidates = _find_image_candidates_on_web(q1, limit=6)
            logger.info(f'[AutoPic] Estratégia 1 retornou {len(candidates)} candidato(s)')

        # Estratégia 2: nome entre aspas + cargo (LLM) + empresa — mais precisa quando há namesakes
        if len(candidates) < 3 and name and company and role:
            q2 = f'"{name}" {role} {company}'
            queries_tried.append(q2)
            logger.info(f'[AutoPic] Estratégia 2 (nome aspas + cargo LLM + empresa): query={q2!r}')
            extra = _find_image_candidates_on_web(q2, limit=6)
            logger.info(f'[AutoPic] Estratégia 2 retornou {len(extra)} candidato(s)')
            candidates = _merge_candidates(candidates, extra)

        # Estratégia 3: nome + empresa sem aspas (mais abrangente, pode trazer mais resultados)
        if len(candidates) < 3 and name and company:
            q3 = f'{name} {company}'
            queries_tried.append(q3)
            logger.info(f'[AutoPic] Estratégia 3 (sem aspas): query={q3!r}')
            extra = _find_image_candidates_on_web(q3, limit=6)
            logger.info(f'[AutoPic] Estratégia 3 retornou {len(extra)} candidato(s)')
            candidates = _merge_candidates(candidates, extra)

        # Estratégia 4: apenas nome (último recurso)
        if len(candidates) < 3:
            q4 = name or company
            queries_tried.append(q4)
            logger.info(f'[AutoPic] Estratégia 4 (apenas nome): query={q4!r}')
            extra = _find_image_candidates_on_web(q4, limit=6)
            logger.info(f'[AutoPic] Estratégia 4 retornou {len(extra)} candidato(s)')
            candidates = _merge_candidates(candidates, extra)

        logger.info(f'[AutoPic] Total final: {len(candidates)} candidato(s) após {len(queries_tried)} estratégia(s). '
                    f'Queries tentadas: {queries_tried}. Cargo LLM: {role!r}')
        return jsonify({'query': queries_tried[0] if queries_tried else '', 'candidates': candidates, 'role': role})
    except Exception as e:
        logger.exception(f'[AutoPic] ERRO em GET /autofind-photo-candidates: {e}')
        return jsonify({'error': f'Erro ao buscar imagens: {e}'}), 500


@app.route('/api/clients/autofind-photo', methods=['POST'])
def clients_autofind_photo():
    try:
        data = request.get_json() or {}
        image_url = (data.get('image_url') or '').strip()
        person_name = (data.get('name') or '').strip() or 'autofind'
        logger.info(f'[AutoPic] POST /autofind-photo: person_name={person_name!r} image_url={image_url[:120]!r}')
        if not image_url:
            return jsonify({'error': 'URL da imagem não informada.'}), 400

        safe_prefix = secure_filename(person_name) or 'autofind'
        photo_url = _download_remote_image_to_uploads(image_url, prefix=safe_prefix)
        logger.info(f'[AutoPic] POST /autofind-photo: imagem salva com sucesso em {photo_url!r}')
        return jsonify({'photo_url': photo_url})
    except Exception as e:
        logger.exception(f'[AutoPic] ERRO em POST /autofind-photo: {type(e).__name__}: {e}')
        return jsonify({'error': f'Erro ao importar imagem selecionada: {e}'}), 500


@app.route('/api/clients/autofind-photo-base64', methods=['POST'])
def clients_autofind_photo_base64():
    """Recebe uma imagem em base64 (data URL), salva no servidor e retorna a URL local."""
    try:
        data = request.get_json() or {}
        data_url = (data.get('data_url') or '').strip()
        person_name = (data.get('name') or '').strip() or 'autofind'
        if not data_url:
            return jsonify({'error': 'data_url não informada.'}), 400

        # Formato esperado: data:image/jpeg;base64,<dados>
        if not data_url.startswith('data:image/'):
            return jsonify({'error': 'Formato de data URL inválido.'}), 400

        header, encoded = data_url.split(',', 1)
        # Extrair extensão do tipo MIME
        mime_type = header.split(';')[0].replace('data:', '')
        ext_map = {'image/jpeg': '.jpg', 'image/png': '.png', 'image/webp': '.webp', 'image/gif': '.gif'}
        ext = ext_map.get(mime_type, '.jpg')

        img_data = base64.b64decode(encoded)
        if len(img_data) > 6 * 1024 * 1024:
            return jsonify({'error': 'Imagem muito grande (máximo 6MB).'}), 400

        safe_prefix = secure_filename(person_name) or 'autofind'
        filename = secure_filename(f"{safe_prefix}-{int(time.time()*1000)}{ext}")
        path = UPLOAD_DIR / filename
        with open(path, 'wb') as f:
            f.write(img_data)

        return jsonify({'photo_url': f'/uploads/{filename}'})
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/clients/autofind-photo-base64: {e}')
        return jsonify({'error': f'Erro ao salvar imagem: {e}'}), 500


@app.route('/api/clientes', methods=['POST'])
def create_cliente():
    return create_client()


@app.route('/api/clients', methods=['POST'])
def create_client():
    try:
        logger.debug('[DEBUG] POST /api/clients')
        logger.debug(f'[DEBUG] Content-Type: {request.content_type}')
        logger.debug(f'[DEBUG] Form data: {request.form}')
        logger.debug(f'[DEBUG] Files: {request.files}')
        
        name = request.form.get('name', '').strip()
        company = request.form.get('company', '').strip()
        position = request.form.get('position', '').strip()
        email = request.form.get('email', '').strip()
        phone = normalize_phone(request.form.get('phone', '').strip())
        linkedin = request.form.get('linkedin', '').strip()
        area_of_activity = request.form.get('area_of_activity', '').strip()
        is_cold_contact = 1 if request.form.get('is_cold_contact') in ('1', 'true', 'on') else 0
        is_target = 1 if request.form.get('is_target') in ('1', 'true', 'on') else 0
        force_create = request.form.get('force_create') in ('1', 'true', 'on')
        autofind_photo_url = (request.form.get('autofind_photo_url') or '').strip()
        
        if not name or not company or not position:
            return jsonify({'error': 'Nome, empresa e cargo sao obrigatorios'}), 400
        
        if not force_create:
            duplicate_clauses = []
            duplicate_params = []
            if name:
                duplicate_clauses.append('LOWER(TRIM(name)) = LOWER(TRIM(?))')
                duplicate_params.append(name)
            if email:
                duplicate_clauses.append('LOWER(TRIM(email)) = LOWER(TRIM(?))')
                duplicate_params.append(email)
            if phone:
                duplicate_clauses.append('TRIM(phone) = TRIM(?)')
                duplicate_params.append(phone)
            if duplicate_clauses:
                conn = get_db()
                c = conn.cursor()
                c.execute(f"SELECT * FROM clients WHERE {' OR '.join([f'({cl})' for cl in duplicate_clauses])} ORDER BY id DESC LIMIT 1", duplicate_params)
                existing = dict_from_row(c.fetchone())
                conn.close()
                if existing:
                    return jsonify({'error': 'Possível duplicidade encontrada', 'duplicate': existing}), 409

        # Suporte a foto em base64 (enviada pela extensão AutoToca)
        if autofind_photo_url and autofind_photo_url.startswith('data:image/'):
            saved = _save_photo_from_base64(autofind_photo_url, person_name=name or 'linkedin')
            photo_url = saved or None
        else:
            photo_url = autofind_photo_url or None
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename:
                filename = secure_filename(file.filename)
                filepath = UPLOAD_DIR / filename
                file.save(str(filepath))
                photo_url = f'/uploads/{filename}'

        conn = get_db()
        c = conn.cursor()
        c.execute('''INSERT INTO clients (name, company, position, area_of_activity, email, phone, linkedin, photo_url, is_target, is_cold_contact, is_archived)
                     VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0)''',
                  (name, company, position, area_of_activity or None, email or None, phone or None, linkedin or None, photo_url, is_target, is_cold_contact))
        client_id = c.lastrowid
        ensure_account_for_company(c, company)
        conn.commit()
        conn.close()
        
        logger.debug(f'[DEBUG] Cliente criado com ID: {client_id}')
        return jsonify({'id': client_id, 'message': 'Cliente criado'}), 201
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/clients: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/clientes/<int:client_id>', methods=['PUT'])
def update_cliente(client_id):
    return update_client(client_id)


@app.route('/api/clients/<int:client_id>', methods=['PUT'])
def update_client(client_id):
    try:
        logger.debug(f'[DEBUG] PUT /api/clients/{client_id}')
        
        name = request.form.get('name', '').strip()
        company = request.form.get('company', '').strip()
        position = request.form.get('position', '').strip()
        email = request.form.get('email', '').strip()
        phone = normalize_phone(request.form.get('phone', '').strip())
        linkedin = request.form.get('linkedin', '').strip()
        area_of_activity = request.form.get('area_of_activity', '').strip()
        is_cold_contact = 1 if request.form.get('is_cold_contact') in ('1', 'true', 'on') else 0
        remove_photo = request.form.get('remove_photo', '0') == '1'
        autofind_photo_url = (request.form.get('autofind_photo_url') or '').strip()
        is_target = 1 if request.form.get('is_target') in ('1', 'true', 'on') else 0
        
        if not name or not company or not position:
            return jsonify({'error': 'Nome, empresa e cargo sao obrigatorios'}), 400
        
        conn = get_db()
        c = conn.cursor()
        
        # Obter cliente atual
        c.execute('SELECT * FROM clients WHERE id = ?', (client_id,))
        client = dict_from_row(c.fetchone())
        if not client:
            conn.close()
            return jsonify({'error': 'Cliente nao encontrado'}), 404
        
        # Suporte a foto em base64 (enviada pela extensão AutoToca)
        if autofind_photo_url and autofind_photo_url.startswith('data:image/'):
            saved = _save_photo_from_base64(autofind_photo_url, person_name=name or 'linkedin')
            photo_url = None if remove_photo else (saved or client['photo_url'])
        else:
            photo_url = None if remove_photo else (autofind_photo_url or client['photo_url'])
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename:
                filename = secure_filename(file.filename)
                filepath = UPLOAD_DIR / filename
                file.save(str(filepath))
                photo_url = f'/uploads/{filename}'
        
        c.execute('''UPDATE clients SET name = ?, company = ?, position = ?, area_of_activity = ?, email = ?, phone = ?, linkedin = ?, photo_url = ?, is_target = ?, is_cold_contact = ?, is_archived = CASE WHEN TRIM(?) != '' THEN 0 ELSE is_archived END, updated_at = CURRENT_TIMESTAMP
                     WHERE id = ?''',
                  (name, company, position, area_of_activity or None, email or None, phone or None, linkedin or None, photo_url, is_target, is_cold_contact, company, client_id))
        ensure_account_for_company(c, company)
        conn.commit()
        conn.close()
        
        logger.debug(f'[DEBUG] Cliente {client_id} atualizado')
        return jsonify({'message': 'Cliente atualizado'})
    except Exception as e:
        logger.exception(f'[ERROR] PUT /api/clients/{client_id}: {e}')
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/clientes/<int:client_id>', methods=['DELETE'])
def delete_cliente(client_id):
    return delete_client(client_id)


@app.route('/api/clients/<int:client_id>', methods=['DELETE'])
def delete_client(client_id):
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute('DELETE FROM clients WHERE id = ?', (client_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Cliente deletado'})
    except Exception as e:
        logger.exception(f'[ERROR] DELETE /api/clients/{client_id}: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clientes/<int:client_id>/arquivar', methods=['POST'])
def archive_cliente(client_id):
    return archive_client(client_id)


@app.route('/api/clients/<int:client_id>/archive', methods=['POST'])
def archive_client(client_id):
    try:
        payload = request.get_json(silent=True) or {}
        remove_company = bool(payload.get('remove_company'))
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT id, company FROM clients WHERE id = ?', (client_id,))
        existing = dict_from_row(c.fetchone())
        if not existing:
            conn.close()
            return jsonify({'error': 'Cliente nao encontrado'}), 404

        old_company = (existing.get('company') or '').strip()
        if remove_company:
            c.execute('''UPDATE clients
                         SET company = '', is_archived = 1, updated_at = CURRENT_TIMESTAMP
                         WHERE id = ?''', (client_id,))

            c.execute(
                '''INSERT INTO activities (client_id, contact_type, information, description)
                   VALUES (?, ?, ?, ?)''',
                (
                    client_id,
                    'Sistema',
                    f'Usuario arquivado. Empresa removida: {old_company or "-"}',
                    'Registro automático de arquivamento'
                )
            )
        else:
            c.execute('''UPDATE clients
                         SET is_archived = 1, updated_at = CURRENT_TIMESTAMP
                         WHERE id = ?''', (client_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Contato arquivado'})
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/clients/{client_id}/archive: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clientes/<int:client_id>/desarquivar', methods=['POST'])
def unarchive_client(client_id):
    try:
        payload = request.get_json(silent=True) or {}
        new_company = (payload.get('company') or '').strip()
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT id, company FROM clients WHERE id = ?', (client_id,))
        existing = dict_from_row(c.fetchone())
        if not existing:
            conn.close()
            return jsonify({'error': 'Cliente nao encontrado'}), 404

        current_company = (existing.get('company') or '').strip()
        if not current_company and not new_company:
            conn.close()
            return jsonify({'error': 'Informe a conta para desarquivar este contato'}), 400

        if new_company and new_company != current_company:
            c.execute('''UPDATE clients
                         SET company = ?, is_archived = 0, updated_at = CURRENT_TIMESTAMP
                         WHERE id = ?''', (new_company, client_id))
            c.execute(
                '''INSERT INTO activities (client_id, contact_type, information, description)
                   VALUES (?, ?, ?, ?)''',
                (
                    client_id,
                    'Sistema',
                    f'Usuario desarquivado. Empresa atribuida: {new_company}',
                    'Registro automático de desarquivamento'
                )
            )
        else:
            c.execute('''UPDATE clients
                         SET is_archived = 0, updated_at = CURRENT_TIMESTAMP
                         WHERE id = ?''', (client_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Contato desarquivado'})
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/clientes/{client_id}/desarquivar: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/<int:client_id>/target', methods=['PUT'])
def update_client_target(client_id):
    try:
        data = request.get_json() or {}
        is_target = 1 if data.get('is_target') else 0

        conn = get_db()
        c = conn.cursor()
        c.execute('UPDATE clients SET is_target = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?', (is_target, client_id))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Target atualizado'})
    except Exception as e:
        logger.exception(f'[ERROR] PUT /api/clients/{client_id}/target: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clients/target-bulk', methods=['POST'])
def update_target_bulk():
    try:
        data = request.get_json() or {}
        company = (data.get('company') or '').strip()
        position = (data.get('position') or '').strip()

        if not company and not position:
            return jsonify({'error': 'Informe empresa ou cargo'}), 400

        conn = get_db()
        c = conn.cursor()
        where = []
        params = []
        if company:
            where.append('company = ?')
            params.append(company)
        if position:
            where.append('position = ?')
            params.append(position)

        c.execute(f"UPDATE clients SET is_target = 1, updated_at = CURRENT_TIMESTAMP WHERE {' AND '.join(where)}", params)
        affected = c.rowcount
        conn.commit()
        conn.close()
        return jsonify({'message': 'Target em massa aplicado', 'affected': affected})
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/clients/target-bulk: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/export/clientes', methods=['GET'])
def export_clientes():
    try:
        import csv
        from io import StringIO

        allowed_fields = {
            'id': ('id', 'ID'),
            'name': ('name', 'Nome'),
            'company': ('company', 'Empresa'),
            'position': ('position', 'Cargo'),
            'area_of_activity': ('area_of_activity', 'Área de Atuação'),
            'email': ('email', 'Email'),
            'phone': ('phone', 'Telefone'),
            'linkedin': ('linkedin', 'LinkedIn'),
            'photo_url': ('photo_url', 'Foto (URL)'),
            'is_target': ('is_target', 'Contato-Alvo'),
            'is_cold_contact': ('is_cold_contact', 'Contato Frio'),
            'created_at': ('created_at', 'Data de Cadastro'),
            'updated_at': ('updated_at', 'Última Atualização'),
        }

        default_fields = ['id', 'name', 'company', 'position', 'email', 'phone', 'linkedin', 'created_at']
        requested_fields = (request.args.get('fields') or '').strip()

        selected_fields = []
        if requested_fields:
            seen = set()
            for raw_field in requested_fields.split(','):
                field = raw_field.strip()
                if not field or field in seen:
                    continue
                if field in allowed_fields:
                    selected_fields.append(field)
                    seen.add(field)

            if not selected_fields:
                return jsonify({'error': 'Nenhum campo válido foi informado para exportação.'}), 400
        else:
            selected_fields = default_fields

        conn = get_db()
        c = conn.cursor()
        db_fields = ', '.join([allowed_fields[field][0] for field in selected_fields])
        c.execute(f'SELECT {db_fields} FROM clients ORDER BY name')
        rows = c.fetchall()
        conn.close()
        
        # Criar CSV em memoria
        output = StringIO()
        writer = csv.writer(output)
        writer.writerow([allowed_fields[field][1] for field in selected_fields])

        for row in rows:
            writer.writerow([row[field] if row[field] is not None else '' for field in selected_fields])
        
        # Retornar como arquivo
        from flask import Response
        response = Response(output.getvalue(), mimetype='text/csv')
        response.headers['Content-Disposition'] = 'attachment; filename=clientes.csv'
        return response
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/export/clientes: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/importar-clientes/template-xlsx', methods=['GET'])
def download_import_clients_template():
    try:
        if not OPENPYXL_AVAILABLE:
            return jsonify({'error': 'Geração de template XLSX requer openpyxl instalado'}), 500

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clientes'
        headers = ['Nome', 'Empresa', 'Cargo', 'Email', 'Telefone', 'Linkedin']
        ws.append(headers)
        header_fill = PatternFill(start_color='34D399', end_color='34D399', fill_type='solid')
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        for col_letter, width in zip('ABCDEF', [25, 25, 20, 30, 18, 30]):
            ws.column_dimensions[col_letter].width = width

        ws.append(['João Silva', 'Tech Corp', 'Gerente', 'joao@techcorp.com', '11999999999', ''])
        ws.append(['Maria Santos', 'Inovação Ltd', 'Diretora', 'maria@inovacao.com', '21988888888', ''])

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name='template_clientes.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/importar-clientes/template-xlsx: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/importar-clientes', methods=['POST'])
def import_clients():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
        
        import csv
        import io
        import tempfile
        
        filename = file.filename.lower()
        rows = []
        
        # VALIDACAO 1: Verificar tamanho do arquivo (maximo 5MB)
        file.seek(0, 2)
        file_size = file.tell()
        file.seek(0)
        
        MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
        if file_size > MAX_FILE_SIZE:
            return jsonify({'error': f'Arquivo muito grande. Maximo: 5MB. Tamanho: {file_size / 1024 / 1024:.2f}MB'}), 400
        
        # Detectar tipo de arquivo
        if filename.endswith('.xls') and not filename.endswith('.xlsx'):
            return jsonify({'error': 'Formato .xls não suportado. Salve como .xlsx ou .csv.'}), 400

        if filename.endswith('.xlsx'):
            try:
                if OPENPYXL_AVAILABLE:
                    from openpyxl import load_workbook

                    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                        file.save(tmp.name)
                        tmp_path = tmp.name

                    try:
                        wb = load_workbook(tmp_path, data_only=True)
                        ws = wb.active

                        for idx, row in enumerate(ws.iter_rows(values_only=True)):
                            if idx == 0:
                                continue
                            row_data = []
                            for cell in row:
                                if cell is None:
                                    row_data.append('')
                                else:
                                    val = str(cell).strip()
                                    if any(ord(c) > 127 and ord(c) < 160 for c in val):
                                        val = val.encode('utf-8', errors='ignore').decode('utf-8')
                                    row_data.append(val)
                            rows.append(row_data)

                        wb.close()
                    finally:
                        os.unlink(tmp_path)
                else:
                    parsed_rows = parse_xlsx_without_openpyxl(file)
                    for idx, row in enumerate(parsed_rows):
                        if idx == 0:
                            continue
                        rows.append(row)
            except Exception as e:
                logger.exception(f'[ERROR] Excel parsing: {e}')
                return jsonify({'error': f'Erro ao ler Excel: {str(e)}'}), 400
        else:
            try:
                # Ler arquivo e detectar encoding
                file_content = file.stream.read()

                # Detectar encoding com chardet quando disponivel,
                # mas manter fallback para nao quebrar importacao.
                if CHARDET_AVAILABLE:
                    detected = chardet.detect(file_content)
                    encoding = detected.get('encoding', 'utf-8') or 'utf-8'
                    try:
                        text = file_content.decode(encoding)
                    except:
                        text = file_content.decode('utf-8', errors='ignore')
                else:
                    # fallback sem dependencia externa
                    try:
                        text = file_content.decode('utf-8')
                    except UnicodeDecodeError:
                        text = file_content.decode('latin-1', errors='ignore')
                
                stream = io.StringIO(text, newline=None)
                csv_data = csv.reader(stream)
                
                for idx, row in enumerate(csv_data):
                    if idx == 0:
                        continue
                    rows.append(row)
            except Exception as e:
                logger.exception(f'[ERROR] CSV parsing: {e}')
                return jsonify({'error': f'Erro ao ler CSV: {str(e)}'}), 400
        
        # VALIDACAO 2: Verificar quantidade de linhas (maximo 1000)
        MAX_ROWS = 1000
        if len(rows) > MAX_ROWS:
            return jsonify({'error': f'Arquivo contem muitas linhas. Maximo: {MAX_ROWS}. Enviado: {len(rows)}'}), 400
        
        if len(rows) == 0:
            return jsonify({'error': 'Arquivo vazio ou sem dados validos'}), 400
        
        # VALIDACAO 3: Validar estrutura dos dados
        invalid_rows = []
        valid_rows = []
        
        for idx, row in enumerate(rows):
            if len(row) < 3:
                invalid_rows.append(f'Linha {idx + 2}: Dados incompletos (minimo 3 campos)')
                continue
            
            name = row[0].strip() if len(row) > 0 else ''
            company = row[1].strip() if len(row) > 1 else ''
            position = row[2].strip() if len(row) > 2 else ''
            email = row[3].strip() if len(row) > 3 else None
            phone = row[4].strip() if len(row) > 4 else None
            linkedin = row[5].strip() if len(row) > 5 else None
            
            # VALIDACAO 4: Campos obrigatorios nao podem estar vazios
            if not name or not company or not position:
                invalid_rows.append(f'Linha {idx + 2}: Nome, Empresa ou Cargo vazios')
                continue
            
            # VALIDACAO 5: Validar tamanho dos campos
            if len(name) > 100 or len(company) > 100 or len(position) > 100:
                invalid_rows.append(f'Linha {idx + 2}: Campos muito longos (maximo 100 caracteres)')
                continue
            
            # VALIDACAO 6: Validar email se fornecido
            if email and '@' not in email:
                invalid_rows.append(f'Linha {idx + 2}: Email invalido')
                continue
            
            valid_rows.append({
                'name': name,
                'company': company,
                'position': position,
                'email': email,
                'phone': phone,
                'linkedin': linkedin
            })
        
        # Se houver erros de validacao, retornar sem importar nada
        if invalid_rows:
            error_msg = 'Erros encontrados no arquivo:\n' + '\n'.join(invalid_rows[:10])
            if len(invalid_rows) > 10:
                error_msg += f'\n... e mais {len(invalid_rows) - 10} erros'
            return jsonify({'error': error_msg}), 400
        
        # IMPORTACAO: Usar transacao para garantir consistencia
        conn = sqlite3.connect(str(DB_PATH))
        c = conn.cursor()
        
        try:
            imported_count = 0
            duplicates_count = 0
            
            for row_data in valid_rows:
                # Verificar se cliente ja existe
                c.execute('SELECT id FROM clients WHERE name = ? AND company = ?', 
                         (row_data['name'], row_data['company']))
                if c.fetchone():
                    duplicates_count += 1
                    continue
                
                # Inserir novo cliente
                c.execute('''INSERT INTO clients (name, company, position, email, phone, linkedin, created_at, updated_at)
                            VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)''',
                         (row_data['name'], row_data['company'], row_data['position'], 
                          row_data['email'], row_data['phone'], row_data.get('linkedin')))
                ensure_account_for_company(c, row_data['company'])
                imported_count += 1
            
            conn.commit()
            
            msg = f'Importados {imported_count} clientes'
            if duplicates_count > 0:
                msg += f'. {duplicates_count} duplicatas ignoradas'
            
            return jsonify({'imported': imported_count, 'duplicates': duplicates_count, 'message': msg}), 200
        
        except Exception as e:
            conn.rollback()
            logger.exception(f'[ERROR] Import transaction failed: {e}')
            return jsonify({'error': f'Erro ao importar dados: {str(e)}'}), 500
        finally:
            conn.close()
    
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/importar-clientes: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/clientes/linkedin-autofill', methods=['POST'])
def client_linkedin_autofill():
    """Preenche automaticamente o cadastro de contato a partir de um perfil LinkedIn via IA."""
    try:
        data = request.get_json() or {}
        linkedin_url = (data.get('linkedin_url') or '').strip()
        profile_text = (data.get('profile_text') or '').strip()
        extension_photo_url = (data.get('extension_photo_url') or '').strip()
        if not linkedin_url and not profile_text:
            return jsonify({'error': 'Cole o texto do perfil LinkedIn ou use a extensão para importar.'}), 400
        task_id = uuid.uuid4().hex
        _bg_task_set(task_id, {'status': 'processing', 'step': 'Iniciando...', 'progress': 5})
        threading.Thread(
            target=_client_linkedin_autofill_async,
            args=(task_id, linkedin_url, profile_text, extension_photo_url),
            daemon=True
        ).start()
        return jsonify({'task_id': task_id}), 202
    except Exception as e:
        logger.exception(f'[ClientLinkedInAutoFill] Erro: {e}')
        return jsonify({'error': str(e)}), 500
