# -*- coding: utf-8 -*-
# Rotas do domínio "wikitoca" (Bloco 3 — modularização).
# Este arquivo é executado no namespace de app.py por _load_route_modules():
# tem acesso a todos os helpers/globals de app.py e registra as rotas no
# mesmo objeto Flask `app`, com URLs idênticas às originais.

# Tabelas de destino permitidas em _wiki_index_document — mapa explícito em vez
# de um catch-all `else`: uma chave desconhecida levanta em vez de gravar em
# silêncio na tabela vizinha (ex.: um typo futuro em `table`).
_WIKI_INDEX_UPDATE_SQL = {
    'wiki_documents': ('UPDATE wiki_documents SET extracted_text=?, extract_status=?, '
                       'extracted_at=CURRENT_TIMESTAMP WHERE id=?'),
    'wiki_training_documents': ('UPDATE wiki_training_documents SET extracted_text=?, '
                                'extract_status=? WHERE id=?'),
}

# Threads de indexação em background disparadas por esta rota (upload, import-zip
# e reindex). Os testes dão join nelas no teardown (ver tests/conftest.py): sem
# isso, uma thread daemon ainda viva sobrevive ao monkeypatch de DB_PATH revertido
# e grava no banco real do usuário com ids do banco de teste.
_wiki_indexing_threads = []


def _wiki_track_thread(thread):
    """Registra uma thread de indexação recém-disparada, descartando da lista as
    que já terminaram — para ela não crescer indefinidamente ao longo da vida do
    processo em produção."""
    global _wiki_indexing_threads
    _wiki_indexing_threads = [t for t in _wiki_indexing_threads if t.is_alive()]
    _wiki_indexing_threads.append(thread)


def _wiki_index_document(table, row_id, file_path):
    """Extrai o texto de um arquivo e grava no cache da linha indicada.
    `table` é 'wiki_documents' ou 'wiki_training_documents'.
    Nunca levanta: falha vira extract_status='error' para aparecer na UI."""
    file_path = Path(file_path)
    if not file_path.exists():
        # Caso realista e recuperável: o arquivo sumiu do disco (limpeza manual,
        # migração incompleta, restore parcial...). `_itoca_extract_text_from_file`
        # também detecta isso e devolve '' em silêncio, mas aqui precisamos
        # distinguir "arquivo existe e não tem texto" (empty) de "arquivo ausente"
        # (error) — senão o selo de erro da UI nunca acende.
        logger.warning(f'[WikiToca] Arquivo não encontrado no disco ao indexar '
                       f'({table} id={row_id}): {file_path}')
        status, texto = 'error', ''
    else:
        try:
            texto = _itoca_extract_text_from_file(str(file_path)) or ''
            status = 'ok' if texto.strip() else 'empty'
        except Exception as e:
            logger.warning(f'[WikiToca] Falha ao extrair texto de {file_path}: {e}')
            status, texto = 'error', ''
    try:
        update_sql = _WIKI_INDEX_UPDATE_SQL.get(table)
        if not update_sql:
            raise ValueError(f'Tabela de indexação desconhecida: {table!r}')
        conn = get_db()
        c = conn.cursor()
        c.execute(update_sql, (texto, status, row_id))
        conn.commit()
        conn.close()
    except Exception as e:
        logger.exception(f'[WikiToca] Falha ao gravar texto extraído ({table} id={row_id}): {e}')
        return 'error'
    if table == 'wiki_documents':
        # Invalidação no lado da ESCRITA para o cache de tokenização da cascata
        # da Capacitação (routes/wikitoca_capacitacao.py). A assinatura por
        # versão que aquele cache usa fecha adição, exclusão e mudança de
        # status sozinha, mas não fecha ESTE caso: reindexar um documento no
        # mesmo segundo em que outro foi indexado deixa COUNT, MAX(id) e
        # MAX(extracted_at) idênticos — `CURRENT_TIMESTAMP` do SQLite tem
        # granularidade de segundo — e a cascata seguiria respondendo com o
        # texto antigo. Esta é a única função que altera `extracted_text` de
        # `wiki_documents`, então avisar daqui fecha o caso por completo.
        _wiki_cap_invalida_cache_da_base()
    return status


def _wiki_index_documents_async(task_id, doc_ids):
    """Indexa uma lista de wiki_documents em background, reportando progresso."""
    try:
        total = len(doc_ids)
        if not total:
            _bg_task_set(task_id, {'status': 'done', 'step': 'Nada a indexar.',
                                   'progress': 100, 'result': {'indexed': 0, 'total': 0}})
            return
        indexados = 0
        for pos, doc_id in enumerate(doc_ids, start=1):
            conn = get_db()
            row = dict_from_row(conn.execute(
                'SELECT file_name, original_name FROM wiki_documents WHERE id=?', (doc_id,)).fetchone())
            conn.close()
            if not row:
                continue
            nome = row.get('original_name') or row.get('file_name')
            _bg_task_set(task_id, {
                'step': f'Processando {pos} de {total} — {nome}',
                'progress': int(5 + (pos - 1) * 90 / total),
            })
            if _wiki_index_document('wiki_documents', doc_id, WIKI_UPLOAD_DIR / row['file_name']) == 'ok':
                indexados += 1
        _bg_task_set(task_id, {'status': 'done', 'step': 'Concluído!', 'progress': 100,
                               'result': {'indexed': indexados, 'total': total}})
    except Exception as e:
        logger.exception(f'[WikiToca] _wiki_index_documents_async: {e}')
        _bg_task_set(task_id, {'status': 'error', 'error': str(e), 'progress': 100})
    finally:
        _bg_task_cleanup(task_id)


# Guarda contra reindexações concorrentes: um duplo-clique no botão "Reindexar"
# (Task 11) não pode disparar duas varreduras de OCR sobre os mesmos ids — não
# corrompe nada, mas duplica o trabalho caro. Só se aplica ao reindex; uploads
# simultâneos continuam legítimos e não passam por este lock.
_wiki_reindex_lock = threading.Lock()
_wiki_reindex_state = {'task_id': None, 'total': 0}


def _wiki_reindex_async(task_id, doc_ids):
    """Roda o backfill de indexação e libera o lock de reindex concorrente ao
    final, com sucesso ou erro."""
    try:
        _wiki_index_documents_async(task_id, doc_ids)
    finally:
        _wiki_reindex_state['task_id'] = None
        _wiki_reindex_lock.release()


@app.route('/api/wikitoca/entries', methods=['GET'])
def list_wiki_entries():
    logger.debug('[DEBUG] GET /api/wikitoca/entries chamado')
    try:
        q = (request.args.get('q') or '').strip()
        conn = get_db()
        c = conn.cursor()
        if q:
            like = f'%{q}%'
            c.execute(
                '''SELECT * FROM wiki_entries
                   WHERE title LIKE ? OR content LIKE ? OR category LIKE ? OR tags LIKE ?
                   ORDER BY updated_at DESC''',
                (like, like, like, like)
            )
        else:
            c.execute('SELECT * FROM wiki_entries ORDER BY updated_at DESC')
        rows = [dict_from_row(r) for r in c.fetchall()]
        conn.close()
        logger.debug(f'[DEBUG] GET /api/wikitoca/entries retornando {len(rows)} registros')
        return jsonify(rows)
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/wikitoca/entries: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_ENTRIES_LIST_ERROR', 'Erro ao listar conhecimentos.', details=str(e),
                         hint='Verifique se o banco de dados está acessível.')


@app.route('/api/wikitoca/entries', methods=['POST'])
def create_wiki_entry():
    logger.debug('[DEBUG] POST /api/wikitoca/entries chamado')
    try:
        data = request.get_json(force=True) or {}
        logger.debug(f'[DEBUG] POST /api/wikitoca/entries payload: {data}')
        title = (data.get('title') or '').strip()
        content = (data.get('content') or '').strip()
        category = (data.get('category') or '').strip() or None
        tags = (data.get('tags') or '').strip() or None
        if not title or not content:
            logger.warning('[WARN] POST /api/wikitoca/entries: titulo ou conteudo ausente')
            return api_error(400, 'WIKI_ENTRY_MISSING_FIELDS', 'Título e conteúdo são obrigatórios.')
        conn = get_db()
        c = conn.cursor()
        c.execute(
            '''INSERT INTO wiki_entries (title, category, content, tags, created_at, updated_at)
               VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)''',
            (title, category, content, tags)
        )
        conn.commit()
        entry_id = c.lastrowid
        c.execute('SELECT * FROM wiki_entries WHERE id = ?', (entry_id,))
        entry = dict_from_row(c.fetchone())
        conn.close()
        logger.debug(f'[DEBUG] POST /api/wikitoca/entries criado id={entry_id}')
        return jsonify(entry), 201
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/wikitoca/entries: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_ENTRY_CREATE_ERROR', 'Erro ao criar conhecimento.', details=str(e))


@app.route('/api/wikitoca/entries/<int:entry_id>', methods=['PUT'])
def update_wiki_entry(entry_id):
    logger.debug(f'[DEBUG] PUT /api/wikitoca/entries/{entry_id} chamado')
    try:
        data = request.get_json(force=True) or {}
        title = (data.get('title') or '').strip()
        content = (data.get('content') or '').strip()
        category = (data.get('category') or '').strip() or None
        tags = (data.get('tags') or '').strip() or None
        if not title or not content:
            return api_error(400, 'WIKI_ENTRY_MISSING_FIELDS', 'Título e conteúdo são obrigatórios.')
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT id FROM wiki_entries WHERE id = ?', (entry_id,))
        if not c.fetchone():
            conn.close()
            logger.warning(f'[WARN] PUT /api/wikitoca/entries/{entry_id}: nao encontrado')
            return api_error(404, 'WIKI_ENTRY_NOT_FOUND', 'Conhecimento não encontrado.')
        c.execute(
            '''UPDATE wiki_entries
               SET title = ?, category = ?, content = ?, tags = ?, updated_at = CURRENT_TIMESTAMP
               WHERE id = ?''',
            (title, category, content, tags, entry_id)
        )
        conn.commit()
        c.execute('SELECT * FROM wiki_entries WHERE id = ?', (entry_id,))
        entry = dict_from_row(c.fetchone())
        conn.close()
        logger.debug(f'[DEBUG] PUT /api/wikitoca/entries/{entry_id} atualizado')
        return jsonify(entry)
    except Exception as e:
        logger.exception(f'[ERROR] PUT /api/wikitoca/entries/{entry_id}: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_ENTRY_UPDATE_ERROR', 'Erro ao atualizar conhecimento.', details=str(e))


@app.route('/api/wikitoca/entries/<int:entry_id>', methods=['DELETE'])
def delete_wiki_entry(entry_id):
    logger.debug(f'[DEBUG] DELETE /api/wikitoca/entries/{entry_id} chamado')
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT id FROM wiki_entries WHERE id = ?', (entry_id,))
        if not c.fetchone():
            conn.close()
            return api_error(404, 'WIKI_ENTRY_NOT_FOUND', 'Conhecimento não encontrado.')
        c.execute('DELETE FROM wiki_entries WHERE id = ?', (entry_id,))
        conn.commit()
        conn.close()
        logger.debug(f'[DEBUG] DELETE /api/wikitoca/entries/{entry_id} removido')
        return jsonify({'message': 'Conhecimento excluído com sucesso.'})
    except Exception as e:
        logger.exception(f'[ERROR] DELETE /api/wikitoca/entries/{entry_id}: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_ENTRY_DELETE_ERROR', 'Erro ao excluir conhecimento.', details=str(e))


# Colunas da listagem de documentos, propositalmente sem `extracted_text`: essa
# rota é chamada a cada troca para a aba WikiToca, e DOCX/XLSX não têm teto de
# tamanho na extração (diferente do PDF, limitado a 30 páginas) — um documento
# grande faria cada troca de aba trafegar dezenas de MB à toa. Quem precisa do
# texto extraído consulta a coluna direto no banco (ex.: busca por conteúdo).
_WIKI_DOC_LIST_COLUMNS = ('id, title, file_name, original_name, file_url, file_ext, file_size, '
                         'extract_status, extracted_at, created_at, updated_at')

_WIKI_EXT_FILTERS = {
    'pdf': ('.pdf',),
    'word': ('.doc', '.docx'),
    'excel': ('.xls', '.xlsx'),
}


def _wiki_norm(texto):
    """Minúsculas e sem acento, para casar 'POLÍTICA' com 'politica'.
    Ignora também caracteres de formatação Unicode (categoria 'Cf' — ex.:
    espaço de largura zero U+200B, hífen suave U+00AD), comuns em texto
    extraído de PDF com hifenização ou quebra de linha automática: sem
    isso, um caractere invisível no meio da palavra quebra um casamento
    que deveria funcionar.

    Atalho ASCII por caractere (mesmo usado em _wiki_norm_indexado, ver lá
    o motivo de ser seguro processar caractere a caractere em vez da string
    inteira de uma vez): pula o normalize() NFKD para a esmagadora maioria
    dos caracteres de um documento real. Medido em português acentuado real
    (contrato de ~480 mil caracteres, ~5% dos caracteres não-ASCII): ~1,4-1,6x
    mais rápido que a versão anterior (NFKD na string inteira de uma vez); o
    ganho cresce com a fração de caracteres ASCII do texto. É o que importa
    aqui — diferente da cascata do ranking (routes/wikitoca_capacitacao.py),
    a busca por conteúdo desta rota paga _wiki_norm por documento a cada
    busca digitada, sem memoização prevista, então o ganho chega direto ao
    usuário. Equivalência com a implementação antiga (NFKD na string inteira)
    é validada em tests/test_wikitoca.py varrendo todo o Unicode."""
    saida = []
    for ch in str(texto or ''):
        if ch.isascii():
            saida.append(ch.lower())
            continue
        for nch in unicodedata.normalize('NFKD', ch):
            if not unicodedata.combining(nch) and unicodedata.category(nch) != 'Cf':
                saida.append(nch.lower())
    return ''.join(saida)


def _wiki_norm_indexado(texto):
    """Como _wiki_norm, mas devolve também o índice do caractere ORIGINAL que
    gerou cada caractere normalizado. Tem que devolver exatamente a mesma
    string que _wiki_norm (mesmos filtros) — senão a posição achada com uma
    função não bate mais com o texto mapeado pela outra.

    NFKD faz decomposição de compatibilidade: 'ﬁ' vira 'fi', '½' vira '1⁄2'.
    Sem esse mapa, uma posição encontrada no texto normalizado aponta para o
    caractere errado no texto original — e o destaque sai deslocado. Ligaduras
    são comuns em texto extraído de PDF, então isso acontece de verdade.

    Atalho ASCII: a esmagadora maioria dos caracteres de um documento é ASCII
    puro (sempre 1 code point, nunca combinante, nunca 'Cf') e não precisa
    passar pelo normalize() caractere a caractere — é o que faz a diferença
    de performance em documentos grandes.
    """
    saida, indices = [], []
    for i, ch in enumerate(str(texto or '')):
        if ch.isascii():
            saida.append(ch.lower())
            indices.append(i)
            continue
        for nch in unicodedata.normalize('NFKD', ch):
            if unicodedata.combining(nch) or unicodedata.category(nch) == 'Cf':
                continue
            saida.append(nch.lower())
            indices.append(i)
    return ''.join(saida), indices


def _wiki_snippet(texto, termo, janela=200):
    """Trecho em volta da primeira ocorrência do termo, com <mark> no termo.
    Tudo que veio do arquivo é escapado; só o <mark> é inserido por nós, em
    posição conhecida — é isso que permite o frontend renderizar sem escapar de
    novo. Devolve '' se o termo não aparecer no texto.

    O trecho dentro do próprio <mark> também é limitado a `janela` caracteres
    do texto ORIGINAL: caracteres combinantes/de formatação entre dois
    caracteres normalizados adjacentes (ex.: um acúmulo patológico de acentos
    combinantes, algo que acontece de verdade em extração malformada de PDF)
    somem do texto normalizado mas continuam ocupando espaço no original —
    sem esse teto, um match de poucos caracteres normalizados pode
    corresponder a dezenas de milhares de caracteres reais, e todo esse
    tamanho vai para innerHTML na listagem."""
    if not texto or not termo:
        return ''
    texto = str(texto)
    termo_norm = _wiki_norm(termo)
    # Um termo formado só por caracteres combinantes/de formatação normaliza
    # para string vazia; `find('')` devolveria 0 e o mapa de índices seria
    # indexado fora do range (IndexError = 500 na rota de busca). Sem termo,
    # não há o que destacar.
    if not termo_norm:
        return ''
    # Curto-circuito: só constrói o mapa de índices (a parte cara, porque
    # passa o NFKD caractere a caractere para todo caractere não-ASCII) se já
    # se sabe que há match — a maioria das buscas não bate na maioria dos
    # documentos, e sem isso a rota fica inviável com dezenas de documentos.
    if termo_norm not in _wiki_norm(texto):
        return ''
    norm, indices = _wiki_norm_indexado(texto)
    pos_norm = norm.find(termo_norm)
    if pos_norm < 0:
        return ''
    termo_len_norm = len(termo_norm)
    # Converte a posição achada no texto normalizado de volta para índices do
    # texto original via o mapa caractere a caractere de _wiki_norm_indexado.
    pos = indices[pos_norm]
    fim_norm = pos_norm + termo_len_norm - 1
    # fim_norm é sempre um índice válido em `indices`: o termo foi encontrado
    # dentro de `norm`, que tem exatamente o mesmo tamanho de `indices`.
    pos_fim = min(indices[fim_norm] + 1, pos + janela)
    ini = max(0, pos - janela // 2)
    fim = min(len(texto), pos_fim + janela // 2)
    antes = html.escape(texto[ini:pos])
    match = html.escape(texto[pos:pos_fim])
    depois = html.escape(texto[pos_fim:fim])
    prefixo = '…' if ini > 0 else ''
    sufixo = '…' if fim < len(texto) else ''
    snippet = f'{prefixo}{antes}<mark>{match}</mark>{depois}{sufixo}'
    return re.sub(r'\s+', ' ', snippet)


@app.route('/api/wikitoca/documents', methods=['GET'])
def list_wiki_documents():
    logger.debug('[DEBUG] GET /api/wikitoca/documents chamado')
    try:
        # Teto defensivo independente do teto de `janela` em _wiki_snippet: um
        # `q` gigantesco não pode virar trabalho proporcional ao tamanho dele.
        q = (request.args.get('q') or '').strip()[:200]
        ext_filtro = (request.args.get('ext') or '').strip().lower()
        conn = get_db()
        c = conn.cursor()
        # Sem busca, o texto extraído nem sai do banco: um DOCX/XLSX grande gera
        # dezenas de MB e esta rota é chamada a cada troca de aba. Só a busca
        # precisa do texto, e ainda assim ele não volta na resposta.
        colunas = _WIKI_DOC_LIST_COLUMNS + (', extracted_text' if q else '')
        # Não filtramos por extract_status aqui de propósito: um documento
        # 'error'/'empty' tem extracted_text vazio e só aparece se o NOME
        # bater, e um 'pending' pode ter texto de uma indexação anterior (por
        # exemplo, reimportado) e continua pesquisável por conteúdo enquanto a
        # nova indexação não termina. Um `WHERE extract_status = 'ok'`
        # pareceria uma otimização inofensiva, mas faria documentos
        # desaparecerem também da busca por NOME.
        c.execute(f'SELECT {colunas} FROM wiki_documents ORDER BY updated_at DESC')
        rows = [dict_from_row(r) for r in c.fetchall()]
        conn.close()

        if ext_filtro in _WIKI_EXT_FILTERS:
            aceitos = _WIKI_EXT_FILTERS[ext_filtro]
            rows = [r for r in rows if (r.get('file_ext') or '').lower() in aceitos]

        if q:
            alvo = _wiki_norm(q)
            filtrados = []
            for r in rows:
                em_nome = alvo in _wiki_norm(r.get('original_name')) or alvo in _wiki_norm(r.get('title'))
                snippet = _wiki_snippet(r.get('extracted_text'), q)
                if em_nome or snippet:
                    r['snippet'] = snippet
                    filtrados.append(r)
            rows = filtrados
        else:
            for r in rows:
                r['snippet'] = ''

        # O texto só foi buscado para calcular o snippet; a UI nunca usa o
        # conteúdo integral, então ele não volta na resposta.
        for r in rows:
            r.pop('extracted_text', None)

        logger.debug(f'[DEBUG] GET /api/wikitoca/documents retornando {len(rows)} documentos')
        return jsonify(rows)
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/wikitoca/documents: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_DOCS_LIST_ERROR', 'Erro ao listar documentos.', details=str(e))


@app.route('/api/wikitoca/documents', methods=['POST'])
def upload_wiki_documents():
    logger.debug('[DEBUG] POST /api/wikitoca/documents chamado')
    try:
        files = request.files.getlist('files')
        logger.debug(f'[DEBUG] POST /api/wikitoca/documents arquivos recebidos: {[f.filename for f in files]}')
        if not files or all(not f.filename for f in files):
            return api_error(400, 'WIKI_DOC_NO_FILE', 'Nenhum arquivo enviado.')
        title = (request.form.get('title') or '').strip()
        conn = get_db()
        c = conn.cursor()
        created = []
        for f in files:
            if not f.filename:
                continue
            ext = Path(f.filename).suffix.lower()
            if ext not in ALLOWED_WIKI_EXTENSIONS:
                logger.warning(f'[WARN] POST /api/wikitoca/documents: extensao rejeitada: {ext}')
                continue
            original_name = f.filename
            safe_name = secure_filename(f'wiki_{int(datetime.now().timestamp())}_{original_name}')
            save_path = WIKI_UPLOAD_DIR / safe_name
            f.save(str(save_path))
            file_size = save_path.stat().st_size
            file_url = f'/uploads/wikitoca/{safe_name}'
            doc_title = title or original_name
            c.execute(
                '''INSERT INTO wiki_documents (title, file_name, original_name, file_url, file_ext, file_size,
                                              extract_status, created_at, updated_at)
                   VALUES (?, ?, ?, ?, ?, ?, 'pending', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)''',
                (doc_title, safe_name, original_name, file_url, ext, file_size)
            )
            conn.commit()
            doc_id = c.lastrowid
            # Mesma invariante da listagem: a resposta do upload não pode
            # trazer o texto extraído (aqui ainda é sempre NULL, só por causa
            # do timing da indexação assíncrona -- não é para depender disso).
            c.execute(f'SELECT {_WIKI_DOC_LIST_COLUMNS} FROM wiki_documents WHERE id = ?', (doc_id,))
            created.append(dict_from_row(c.fetchone()))
            logger.debug(f'[DEBUG] POST /api/wikitoca/documents salvo id={doc_id} nome={original_name}')
        conn.close()
        if not created:
            return api_error(400, 'WIKI_DOC_INVALID_TYPE',
                             'Nenhum arquivo válido enviado. Tipos aceitos: PDF, XLS, XLSX, DOC, DOCX.')
        task_id = uuid.uuid4().hex
        _bg_task_register_persistent(task_id, 'wiki_indexacao')
        _bg_task_set(task_id, {'status': 'processing', 'step': 'Indexando documentos...', 'progress': 5})
        thread = threading.Thread(target=_wiki_index_documents_async,
                                  args=(task_id, [d['id'] for d in created]), daemon=True)
        _wiki_track_thread(thread)
        thread.start()
        return jsonify({'documents': created, 'task_id': task_id}), 201
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/wikitoca/documents: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_DOC_UPLOAD_ERROR', 'Erro ao enviar documento.', details=str(e))


@app.route('/api/wikitoca/documents/reindex', methods=['POST'])
def reindex_wiki_documents():
    """Backfill do texto extraído dos documentos já existentes.
    Body opcional: {"force": true} para reprocessar também os já indexados."""
    logger.debug('[DEBUG] POST /api/wikitoca/documents/reindex chamado')
    if not _wiki_reindex_lock.acquire(blocking=False):
        logger.info(f"[WikiToca] Reindexação já em andamento (task_id={_wiki_reindex_state['task_id']}), "
                    "ignorando novo disparo")
        return jsonify({'task_id': _wiki_reindex_state['task_id'],
                        'total': _wiki_reindex_state['total'],
                        'already_running': True}), 202
    try:
        force = bool((request.get_json(silent=True) or {}).get('force'))
        conn = get_db()
        c = conn.cursor()
        if force:
            c.execute('SELECT id FROM wiki_documents ORDER BY id')
        else:
            c.execute("SELECT id FROM wiki_documents "
                      "WHERE extract_status IS NULL OR extract_status IN ('pending', 'error') "
                      "ORDER BY id")
        doc_ids = [r[0] for r in c.fetchall()]
        conn.close()
        task_id = uuid.uuid4().hex
        _wiki_reindex_state['task_id'] = task_id
        _wiki_reindex_state['total'] = len(doc_ids)
        _bg_task_register_persistent(task_id, 'wiki_indexacao')
        _bg_task_set(task_id, {'status': 'processing', 'step': 'Iniciando...', 'progress': 5})
        thread = threading.Thread(target=_wiki_reindex_async, args=(task_id, doc_ids), daemon=True)
        _wiki_track_thread(thread)
        thread.start()
        logger.info(f'[WikiToca] Reindexação iniciada para {len(doc_ids)} documento(s)')
        return jsonify({'task_id': task_id, 'total': len(doc_ids)}), 202
    except Exception as e:
        _wiki_reindex_state['task_id'] = None
        _wiki_reindex_lock.release()
        logger.exception(f'[ERROR] POST /api/wikitoca/documents/reindex: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_DOC_REINDEX_ERROR', 'Erro ao reindexar documentos.', details=str(e))


@app.route('/api/wikitoca/documents/<int:document_id>', methods=['DELETE'])
def delete_wiki_document(document_id):
    logger.debug(f'[DEBUG] DELETE /api/wikitoca/documents/{document_id} chamado')
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT * FROM wiki_documents WHERE id = ?', (document_id,))
        row = dict_from_row(c.fetchone())
        if not row:
            conn.close()
            return api_error(404, 'WIKI_DOC_NOT_FOUND', 'Documento não encontrado.')
        file_path = WIKI_UPLOAD_DIR / row['file_name']
        if file_path.exists():
            file_path.unlink()
        c.execute('DELETE FROM wiki_documents WHERE id = ?', (document_id,))
        conn.commit()
        conn.close()
        logger.debug(f'[DEBUG] DELETE /api/wikitoca/documents/{document_id} removido')
        return jsonify({'message': 'Documento removido com sucesso.'})
    except Exception as e:
        logger.exception(f'[ERROR] DELETE /api/wikitoca/documents/{document_id}: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_DOC_DELETE_ERROR', 'Erro ao remover documento.', details=str(e))


@app.route('/api/wikitoca/documents/export-zip', methods=['GET'])
def export_wiki_documents():
    try:
        from flask import send_file
        import tempfile
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT * FROM wiki_documents ORDER BY id')
        rows = [dict_from_row(r) for r in c.fetchall()]
        conn.close()

        temp_dir = tempfile.mkdtemp()
        temp_zip = Path(temp_dir) / 'wikitoca-documentos.zip'
        manifest = []
        with zipfile.ZipFile(str(temp_zip), mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
            for row in rows:
                file_path = WIKI_UPLOAD_DIR / row['file_name']
                if file_path.exists():
                    zf.write(str(file_path), arcname=f"files/{row['file_name']}")
                manifest.append({
                    'title': row.get('title'),
                    'file_name': row.get('file_name'),
                    'original_name': row.get('original_name'),
                    'file_ext': row.get('file_ext'),
                })
            zf.writestr('manifest.json', json.dumps(manifest, ensure_ascii=False, indent=2))

        return send_file(
            str(temp_zip),
            as_attachment=True,
            download_name=f'wikitoca-documentos-{datetime.now().strftime("%Y%m%d-%H%M%S")}.zip',
            mimetype='application/zip'
        )
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/wikitoca/documents/export-zip: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_DOC_EXPORT_ERROR', 'Erro ao exportar documentos.', details=str(e))


@app.route('/api/wikitoca/documents/import-zip', methods=['POST'])
def import_wiki_documents():
    try:
        if 'file' not in request.files:
            return api_error(400, 'WIKI_DOC_IMPORT_NO_FILE', 'Nenhum arquivo enviado.')
        file = request.files['file']
        if not file.filename or not file.filename.lower().endswith('.zip'):
            return api_error(400, 'WIKI_DOC_IMPORT_INVALID', 'Envie um arquivo .zip exportado pelo Toca do Coelho.')

        import tempfile
        temp_dir = tempfile.mkdtemp()
        temp_zip_path = Path(temp_dir) / 'import.zip'
        file.save(str(temp_zip_path))

        imported = []
        with zipfile.ZipFile(str(temp_zip_path), mode='r') as zf:
            names = zf.namelist()
            if 'manifest.json' not in names:
                return api_error(400, 'WIKI_DOC_IMPORT_INVALID', 'Arquivo .zip inválido: manifest.json não encontrado.')
            manifest = json.loads(zf.read('manifest.json').decode('utf-8'))

            conn = get_db()
            c = conn.cursor()
            for entry in manifest:
                original_name = entry.get('original_name') or entry.get('file_name') or 'documento'
                ext = (entry.get('file_ext') or Path(original_name).suffix.lower())
                if ext not in ALLOWED_WIKI_EXTENSIONS:
                    continue
                src_name = f"files/{entry.get('file_name')}"
                if src_name not in names:
                    continue
                safe_name = secure_filename(f'wiki_{int(datetime.now().timestamp()*1000)}_{original_name}')
                save_path = WIKI_UPLOAD_DIR / safe_name
                with zf.open(src_name) as src, open(save_path, 'wb') as dst:
                    dst.write(src.read())
                file_size = save_path.stat().st_size
                file_url = f'/uploads/wikitoca/{safe_name}'
                doc_title = entry.get('title') or original_name
                c.execute(
                    '''INSERT INTO wiki_documents (title, file_name, original_name, file_url, file_ext, file_size,
                                                  extract_status, created_at, updated_at)
                       VALUES (?, ?, ?, ?, ?, ?, 'pending', CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)''',
                    (doc_title, safe_name, original_name, file_url, ext, file_size)
                )
                conn.commit()
                doc_id = c.lastrowid
                c.execute('SELECT * FROM wiki_documents WHERE id = ?', (doc_id,))
                imported.append(dict_from_row(c.fetchone()))
            conn.close()

        result = {'imported': len(imported), 'documents': imported}
        if imported:
            task_id = uuid.uuid4().hex
            _bg_task_register_persistent(task_id, 'wiki_indexacao')
            _bg_task_set(task_id, {'status': 'processing', 'step': 'Indexando documentos importados...',
                                   'progress': 5})
            thread = threading.Thread(target=_wiki_index_documents_async,
                                      args=(task_id, [d['id'] for d in imported]), daemon=True)
            _wiki_track_thread(thread)
            thread.start()
            result['task_id'] = task_id
        return jsonify(result), 201
    except zipfile.BadZipFile:
        return api_error(400, 'WIKI_DOC_IMPORT_INVALID', 'Arquivo .zip inválido ou corrompido.')
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/wikitoca/documents/import-zip: {e}')
        traceback.print_exc()
        return api_error(500, 'WIKI_DOC_IMPORT_ERROR', 'Erro ao importar documentos.', details=str(e))


@app.route('/api/wikitoca/entries/export-xlsx', methods=['GET'])
def export_wikitoca_xlsx():
    try:
        logger.debug('[DEBUG] GET /api/wikitoca/entries/export-xlsx chamado')
        if not OPENPYXL_AVAILABLE:
            return jsonify({'error': 'Exportação XLSX requer openpyxl instalado'}), 500
        conn = get_db()
        c = conn.cursor()
        c.execute('SELECT title, category, tags, content FROM wiki_entries ORDER BY updated_at DESC')
        rows = c.fetchall()
        conn.close()
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Conhecimentos'
        headers = ['Título', 'Categoria', 'Tags', 'Descrição']
        ws.append(headers)
        header_fill = PatternFill(start_color='34D399', end_color='34D399', fill_type='solid')
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 60
        for row in rows:
            ws.append([
                row['title'] or '',
                row['category'] or '',
                row['tags'] or '',
                row['content'] or ''
            ])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name='wikitoca_conhecimentos.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/wikitoca/entries/export-xlsx: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/wikitoca/entries/template-xlsx', methods=['GET'])
def wikitoca_template_xlsx():
    try:
        logger.debug('[DEBUG] GET /api/wikitoca/entries/template-xlsx chamado')
        if not OPENPYXL_AVAILABLE:
            return jsonify({'error': 'Template XLSX requer openpyxl instalado'}), 500
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Conhecimentos'
        headers = ['Título', 'Categoria', 'Descrição']
        ws.append(headers)
        header_fill = PatternFill(start_color='34D399', end_color='34D399', fill_type='solid')
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 60
        ws.append(['Exemplo de título', 'Comercial', 'Descreva aqui o conhecimento a ser registrado.'])
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name='wikitoca_template.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.exception(f'[ERROR] GET /api/wikitoca/entries/template-xlsx: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/api/wikitoca/entries/import-xlsx', methods=['POST'])
def import_wikitoca_xlsx():
    try:
        logger.debug('[DEBUG] POST /api/wikitoca/entries/import-xlsx chamado')
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        file = request.files['file']
        if not file.filename or not file.filename.lower().endswith('.xlsx'):
            return jsonify({'error': 'Envie um arquivo .xlsx'}), 400
        if not OPENPYXL_AVAILABLE:
            return jsonify({'error': 'Importação XLSX requer openpyxl instalado'}), 500
        import openpyxl
        import tempfile, os as _os
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            rows_data = list(ws.iter_rows(values_only=True))
        finally:
            _os.unlink(tmp_path)
        if not rows_data:
            return jsonify({'error': 'Arquivo vazio'}), 400
        # Detectar colunas pelo cabeçalho
        header = [str(c).strip().lower() if c else '' for c in rows_data[0]]
        def find_col(names):
            for name in names:
                if name in header:
                    return header.index(name)
            return None
        col_title = find_col(['título', 'titulo', 'title'])
        col_cat = find_col(['categoria', 'category'])
        col_desc = find_col(['descrição', 'descricao', 'description', 'conteúdo', 'conteudo', 'content'])
        if col_title is None or col_desc is None:
            return jsonify({'error': 'Colunas obrigatórias não encontradas. O arquivo deve ter colunas Título e Descrição.'}), 400
        # Função de geração de tags (mesma lógica do frontend)
        stopwords = {'a','o','os','as','de','da','do','das','dos','e','é','em','no','na','nos','nas','um','uma','uns','umas','para','por','com','sem','que','se','ao','aos','à','às','ou','como','mais','menos','ja','não','sim'}
        def generate_tags(title, content):
            import re
            text = f'{title or ""} {content or ""}'.lower()
            words = re.findall(r'[a-záàãâéêíóôõúüç0-9-]{3,}', text)
            rank = {}
            for w in words:
                if w in stopwords or w.isdigit():
                    continue
                rank[w] = rank.get(w, 0) + 1
            sorted_words = sorted(rank.items(), key=lambda x: (-x[1], x[0]))
            return ', '.join(w for w, _ in sorted_words[:6])
        conn = get_db()
        c = conn.cursor()
        ok = 0
        fail = 0
        errors = []
        for idx, row in enumerate(rows_data[1:], start=2):
            try:
                title = str(row[col_title]).strip() if row[col_title] else ''
                category = str(row[col_cat]).strip() if col_cat is not None and row[col_cat] else ''
                content = str(row[col_desc]).strip() if row[col_desc] else ''
                if not title or not content:
                    fail += 1
                    errors.append(f'Linha {idx}: título ou descrição vazia')
                    continue
                tags = generate_tags(title, content)
                now = datetime.utcnow().isoformat() + 'Z'
                c.execute(
                    'INSERT INTO wiki_entries (title, category, tags, content, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?)',
                    (title, category, tags, content, now, now)
                )
                ok += 1
            except Exception as row_err:
                fail += 1
                errors.append(f'Linha {idx}: {str(row_err)}')
        conn.commit()
        conn.close()
        logger.debug(f'[DEBUG] POST /api/wikitoca/entries/import-xlsx: {ok} importados, {fail} erros')
        return jsonify({'imported': ok, 'failed': fail, 'errors': errors[:10]}), 200
    except Exception as e:
        logger.exception(f'[ERROR] POST /api/wikitoca/entries/import-xlsx: {e}')
        return jsonify({'error': str(e)}), 500


@app.route('/uploads/wikitoca/<filename>')
def serve_wikitoca_upload(filename):
    return send_from_directory(str(WIKI_UPLOAD_DIR), filename)
